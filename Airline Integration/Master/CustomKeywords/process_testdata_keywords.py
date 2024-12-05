from datetime import datetime

from robot.api.deco import keyword
from openpyxl import load_workbook
import json

@keyword
def fetch_testdata_by_id(file_path, sheet_name, target_id):
    global workbook
    try:
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]
        header = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == target_id:
                data_dict = {}
                for col_name, value in zip(header, row):
                    if ',' in str(value):
                        data_dict[col_name] = [item.strip() for item in value.split(',')]
                    else:
                        data_dict[col_name] = value
                return data_dict
    except Exception as e:
        print(f"Error reading Excel file: {e}")
    finally:
        try:
            workbook.close()
        except NameError:
            pass
    return None

@keyword
def fetch_user_login_data(json_file, user_type, user_id):
    try:
        with open(json_file, 'r') as file:
            data = json.load(file)
            users = data.get('users', {}).get(user_type, [])
            for user in users:
                try:
                    if user.get('user_id') == user_id:
                        user_data = {'Username': user.get('user_name'), 'Password': user.get('password')}
                        return user_data
                except (KeyError, ValueError):
                    print("Error processing user:", user)
                    continue
            raise ValueError("User not found or incorrect user ID")
        print("User not found.")
        return None

    except (FileNotFoundError, json.JSONDecodeError, KeyError) as e:
        raise type(e)(str(e) + f" in JSON file '{json_file}'")

